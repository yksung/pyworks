# -*- coding : utf-8 -*-
'''
Created on 2017. 7. 7.

@author: Holly
'''
from sys import argv
from openpyxl import load_workbook
import xlsxwriter
import os, time
from com.wisenut.dmap_analyzer import get_emotion_type
import sys

if __name__ == '__main__':
    to_read = argv[1]
        
    # 쓸 파일
    excel_to_write = xlsxwriter.Workbook(os.path.join(os.path.dirname(to_read), os.path.basename(to_read[:-5])+"_result.xlsx"))
    
    excel_to_read = load_workbook(to_read, keep_links=False)
    for sheetname in excel_to_read.get_sheet_names():
        w_worksheet = excel_to_write.add_worksheet(sheetname)
        sheet = excel_to_read.get_sheet_by_name(sheetname)
        print(sheet)
        
        print("<%d>"%sheet.max_row)
        for ridx, row in enumerate(sheet.rows):
            if ridx % 100 == 0 and ridx % 1000 != 0:
                print(".", end="")
                sys.stdout.flush()
            if ridx % 1000 == 0:
                time.sleep(3)
                print("[%d]"%ridx, end="")
                sys.stdout.flush()
            if ridx  == 0:
                colnames = { cell.value:idx for idx, cell in enumerate(row) }
                #print(colnames)
                
                w_worksheet.write(0, 0, '매칭문장') # ID
                w_worksheet.write(0, 1, '세분류') # ID
                w_worksheet.write(0, 2, '개수 : 매칭문장') # ID
                w_worksheet.write(0, 3, '긍부정') # ID
            else:
                docid = sheetname + str(ridx)
                sentence = row[colnames['매칭문장']].value
                #print(docid, sentence)
                
                emotion = get_emotion_type(docid, sentence, 1)
                
                w_worksheet.write(ridx, 0, row[colnames['매칭문장']].value) # ID
                w_worksheet.write(ridx, 1, row[colnames['세분류']].value) # ID
                w_worksheet.write(ridx, 2, row[colnames['개수 : 매칭문장']].value) # ID
                w_worksheet.write(ridx, 3, emotion) # ID
                

   
                
    excel_to_write.close()
